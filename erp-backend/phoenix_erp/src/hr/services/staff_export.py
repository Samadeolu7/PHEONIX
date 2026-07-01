# hr/services/staff_export.py
"""
Staff Payroll Excel Export Service
====================================

Generates an Excel (.xlsx) file in the EXACT same layout as the standard
payroll upload template:

  Row 1  – Tenant / school name (merged across all columns)
  Row 2  – "PRIMARY STAFF SALARY FOR <MONTH YEAR>" (merged)
  Row 3  – Main column headers
  Row 4  – Percentages  (16%, 10%, 6% …)
  Row 5  – "Deductions" sub-label merged under deduction columns
  Row 6  – =N= currency row + PAYE PIN / PENSION (PEN number) / PFA / Bank /
            Account Number labels
  Row 7+ – One row per staff member

The column order matches the import template exactly so the downloaded file
can be used directly as an upload after manual edits.
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime
from decimal import Decimal

from django.utils import timezone


# ---------------------------------------------------------------------------
# Column definitions – same order as the import template
# ---------------------------------------------------------------------------

# (header_label, internal_key)
COLUMNS = [
    ('Name',                   'name'),
    ('Basic Salary',           'basic'),
    ('Housing Allowance',      'housing'),
    ('Transport Allowance',    'transport'),
    ('Entertain.',             'entertainment'),
    ('Utility',                'utility'),
    ('Lunch',                  'lunch'),
    ('Leav Allow.',            'leave_allow'),
    ('Gross Salary',           'gross_salary'),
    ('PAYE Deduct',            'paye_deduct'),
    ('Loan Deductions',        'loan_deductions'),
    ('Pension Deductions',     'pension_deductions'),
    ('Dev. Levy & Other',      'dev_levy'),
    ('Other Deductions',       'other_deductions'),
    ('Staff IOU Monthly',      'staff_iou_monthly'),
    ('Staff IOU Balance',      'staff_iou_balance'),
    ('Total Deductions',       'total_deductions'),
    ('Net Pay',                'net_pay'),
    ('PAYE PIN',               'paye_pin'),
    ('PENSION (PEN number)',   'pension_number'),
    ('PFA',                    'pfa'),
    ('Bank',                   'bank_name'),
    ('Account Number',         'bank_account_number'),
]

# Maps SalaryComponent.name → internal key used in COLUMNS above
_COMPONENT_NAME_MAP: dict[str, str] = {
    'basic salary':            'basic',
    'housing allowance':       'housing',
    'transport allowance':     'transport',
    'entertainment allowance': 'entertainment',
    'utility allowance':       'utility',
    'lunch allowance':         'lunch',
    'leave allowance':         'leave_allow',
    'paye tax deduction':      'paye_deduct',
    'loan deduction':          'loan_deductions',
    'pension deduction':       'pension_deductions',
    'development levy':        'dev_levy',
}

# Indices (0-based) of numeric value columns (not name/pin/bank text columns)
_EARNING_KEYS  = {'basic', 'housing', 'transport', 'entertainment', 'utility', 'lunch', 'leave_allow'}
_DEDUCT_KEYS   = {'paye_deduct', 'loan_deductions', 'pension_deductions', 'dev_levy'}

# Percentage hints shown in row 4 (empty string = not shown)
_PERCENT_ROW: dict[str, str] = {
    'basic':          '16%',
    'housing':        '10%',
    'transport':       '6%',
    'entertainment':   '5%',
    'utility':         '3%',
    'lunch':           '3%',
    'leave_allow':     '5%',
}


class StaffPayrollExportService:
    """
    Usage::

        service = StaffPayrollExportService(queryset, period_label="MARCH 2026")
        buffer  = service.generate()           # returns io.BytesIO
    """

    def __init__(self, staff_queryset, period_label: str = ''):
        self.staff_qs      = staff_queryset
        self.period_label  = period_label or self._default_period_label()

    # ------------------------------------------------------------------
    # Public entry point
    # ------------------------------------------------------------------

    def generate(self) -> io.BytesIO:
        """Build the workbook and return it as a BytesIO buffer."""
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, Alignment, PatternFill, Border, Side, numbers
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl is required. Install it: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payroll"

        num_cols   = len(COLUMNS)
        col_keys   = [c[1] for c in COLUMNS]
        col_labels = [c[0] for c in COLUMNS]

        # ── Styles ───────────────────────────────────────────────────────
        header_fill  = PatternFill("solid", fgColor="1F3864")   # dark navy
        header_font  = Font(bold=True, color="FFFFFF", size=10)
        sub_fill     = PatternFill("solid", fgColor="BDD7EE")   # light blue
        pct_fill     = PatternFill("solid", fgColor="D9E1F2")
        deduct_fill  = PatternFill("solid", fgColor="FCE4D6")   # light orange
        earn_fill    = PatternFill("solid", fgColor="E2EFDA")   # light green

        thin   = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right  = Alignment(horizontal='right',  vertical='center')
        left   = Alignment(horizontal='left',   vertical='center')

        # ── Row 1: Tenant/school name ────────────────────────────────────
        ws.append([''] * num_cols)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = self._tenant_name()
        title_cell.font  = Font(bold=True, size=14, color="1F3864")
        title_cell.alignment = center

        # ── Row 2: Month/year heading ────────────────────────────────────
        ws.append([''] * num_cols)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        period_cell = ws.cell(row=2, column=1)
        period_cell.value = f"PRIMARY STAFF SALARY FOR {self.period_label.upper()}"
        period_cell.font  = Font(bold=True, size=12, color="1F3864")
        period_cell.alignment = center

        # ── Row 3: Main headers ──────────────────────────────────────────
        ws.append(col_labels)
        for col_idx, key in enumerate(col_keys, start=1):
            cell = ws.cell(row=3, column=col_idx)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = border

        # ── Row 4: Percentage hints ──────────────────────────────────────
        pct_row = [_PERCENT_ROW.get(k, '') for k in col_keys]
        ws.append(pct_row)
        for col_idx, _ in enumerate(col_keys, start=1):
            cell = ws.cell(row=4, column=col_idx)
            cell.fill      = pct_fill
            cell.alignment = center
            cell.font      = Font(size=9, italic=True, color="595959")
            cell.border    = border

        # ── Row 5: "Deductions" label ────────────────────────────────────
        first_deduct = col_keys.index('paye_deduct') + 1
        last_deduct  = col_keys.index('total_deductions') + 1
        ws.append([''] * num_cols)
        deduct_label_cell = ws.cell(row=5, column=first_deduct)
        deduct_label_cell.value     = 'Deductions'
        deduct_label_cell.font      = Font(bold=True, size=9, color="C00000")
        deduct_label_cell.alignment = center
        deduct_label_cell.fill      = deduct_fill
        deduct_label_cell.border    = border
        if first_deduct < last_deduct:
            ws.merge_cells(
                start_row=5, start_column=first_deduct,
                end_row=5,   end_column=last_deduct,
            )
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=5, column=col_idx)
            cell.border = border
            if col_idx < first_deduct or col_idx > last_deduct:
                cell.fill = pct_fill

        # ── Row 6: =N= / column sub-labels ──────────────────────────────
        n_row = ['=N=' if k in (_EARNING_KEYS | _DEDUCT_KEYS | {'gross_salary', 'net_pay'}) else '' for k in col_keys]
        # Override text columns with their sub-labels
        for label_key, label_text in [
            ('paye_pin',        'PAYE PIN'),
            ('pension_number',  'PENSION (PEN number)'),
            ('fpa',             'PFA'),
            ('bank_name',       'Bank'),
            ('bank_account_number', 'Account Number'),
        ]:
            if label_key in col_keys:
                n_row[col_keys.index(label_key)] = label_text
        n_row[col_keys.index('name')] = '=N='
        ws.append(n_row)
        for col_idx, key in enumerate(col_keys, start=1):
            cell = ws.cell(row=6, column=col_idx)
            cell.alignment = center
            cell.font      = Font(size=9, color="595959")
            cell.border    = border
            if key in _DEDUCT_KEYS or key in ('total_deductions',):
                cell.fill = deduct_fill
            elif key in _EARNING_KEYS or key in ('gross_salary', 'net_pay'):
                cell.fill = earn_fill
            else:
                cell.fill = sub_fill

        # ── Freeze header rows ───────────────────────────────────────────
        ws.freeze_panes = 'A7'

        # ── Data rows ────────────────────────────────────────────────────
        from hr.models import StaffPayInfo, SalaryComponent
        from hr.config_models import HRConfig

        staff_list = list(
            self.staff_qs
            .prefetch_related('pay_info__component')
            .order_by('last_name', 'first_name')
        )

        # Load branch HR config once for PAYE/pension auto-calculation
        hr_config = None
        if staff_list:
            try:
                hr_config = HRConfig.get_for_branch(staff_list[0].branch)
            except Exception:
                pass

        # Resolve the payroll month represented by this export heading so
        # IOU monthly deductions respect each IOU's configured start_month.
        payroll_month = self._resolve_payroll_month(self.period_label)

        # Preload active IOUs for all staff once to avoid N+1 queries.
        ious_by_staff_id: dict[int, list] = {}
        if staff_list:
            from hr.models import StaffIOU

            iou_qs = StaffIOU.objects.filter(
                staff_id__in=[s.id for s in staff_list],
                status=StaffIOU.ACTIVE,
                start_month__lte=payroll_month,
                is_deleted=False,
            ).order_by('staff_id', 'created_at')
            for iou in iou_qs:
                ious_by_staff_id.setdefault(iou.staff_id, []).append(iou)

        number_fmt = '#,##0.00'

        for row_num, staff in enumerate(staff_list, start=7):
            pay_map, taxable_income, pensionable_income, other_deductions_total = self._build_pay_map(staff)

            # Active IOU transparency for this month.
            staff_ious = ious_by_staff_id.get(staff.id, [])
            iou_monthly = sum(
                min(iou.monthly_installment, iou.balance_remaining)
                for iou in staff_ious
            )
            iou_balance = sum(iou.balance_remaining for iou in staff_ious)

            # Calculate gross earnings from earning components
            gross = sum(pay_map.get(k, Decimal('0')) for k in _EARNING_KEYS)

            # Auto-calculate pension on pensionable base (Basic + Housing + Transport)
            # If pensionable_income is 0 (e.g. no is_pensionable flags set), fall back to
            # the known-pensionable keys directly from pay_map.
            pension_base = pensionable_income or sum(
                pay_map.get(k, Decimal('0')) for k in ('basic', 'housing', 'transport')
            )
            if not pay_map.get('pension_deductions') and hr_config and not staff.is_pension_exempt:
                pay_map['pension_deductions'] = hr_config.calculate_employee_pension(pension_base)

            # Auto-calculate PAYE if not set as an explicit pay component
            if not pay_map.get('paye_deduct') and hr_config:
                pension_for_tax = pay_map.get('pension_deductions', Decimal('0'))
                pay_map['paye_deduct'] = hr_config.calculate_tax(taxable_income, pension_for_tax)

            total_deduct = (
                sum(pay_map.get(k, Decimal('0')) for k in _DEDUCT_KEYS)
                + other_deductions_total
                + iou_monthly
            )
            net = gross - total_deduct

            row_data = []
            for key in col_keys:
                if key == 'name':
                    row_data.append(f"{staff.first_name} {staff.last_name}".strip())
                elif key == 'gross_salary':
                    row_data.append(gross)
                elif key == 'other_deductions':
                    row_data.append(other_deductions_total if other_deductions_total else '')
                elif key == 'staff_iou_monthly':
                    row_data.append(iou_monthly if iou_monthly else '')
                elif key == 'staff_iou_balance':
                    row_data.append(iou_balance if iou_balance else '')
                elif key == 'total_deductions':
                    row_data.append(total_deduct)
                elif key == 'net_pay':
                    row_data.append(net)
                elif key == 'paye_pin':
                    row_data.append(staff.paye_pin or '')
                elif key == 'pension_number':
                    row_data.append(staff.pension_number or '')
                elif key == 'fpa':
                    row_data.append(staff.pension_provider or '')
                elif key == 'bank_name':
                    row_data.append(staff.bank_name or '')
                elif key == 'bank_account_number':
                    row_data.append(staff.bank_account_number or '')
                else:
                    val = pay_map.get(key, Decimal('0'))
                    row_data.append(val if val else '')

            ws.append(row_data)

            # Format cells
            row_fill_earn   = PatternFill("solid", fgColor="F2F8EE")
            row_fill_deduct = PatternFill("solid", fgColor="FFF2CC")

            for col_idx, key in enumerate(col_keys, start=1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.border    = border
                cell.alignment = left if key == 'name' else right
                if key in _EARNING_KEYS or key in ('gross_salary',):
                    cell.number_format = number_fmt
                    cell.fill          = row_fill_earn
                elif key in _DEDUCT_KEYS or key in ('other_deductions', 'staff_iou_monthly', 'total_deductions'):
                    cell.number_format = number_fmt
                    cell.fill          = row_fill_deduct
                elif key in ('staff_iou_balance',):
                    cell.number_format = number_fmt
                elif key == 'net_pay':
                    cell.number_format = number_fmt
                    cell.font          = Font(bold=True)

        # ── Totals row ───────────────────────────────────────────────────
        if staff_list:
            total_row_num = 7 + len(staff_list)
            ws.append(['TOTAL'] + [''] * (num_cols - 1))
            total_cell = ws.cell(row=total_row_num, column=1)
            total_cell.font      = Font(bold=True)
            total_cell.alignment = left
            total_cell.fill      = header_fill
            total_cell.font      = Font(bold=True, color='FFFFFF')
            total_cell.border    = border

            for col_idx, key in enumerate(col_keys, start=1):
                if col_idx == 1:
                    continue
                cell = ws.cell(row=total_row_num, column=col_idx)
                cell.border = border
                if key in (_EARNING_KEYS | _DEDUCT_KEYS | {'other_deductions', 'staff_iou_monthly', 'gross_salary', 'total_deductions', 'net_pay'}):
                    # SUM formula over the data rows
                    col_letter = get_column_letter(col_idx)
                    cell.value         = f"=SUM({col_letter}7:{col_letter}{total_row_num - 1})"
                    cell.number_format = number_fmt
                    cell.font          = Font(bold=True)
                    cell.fill          = header_fill
                    cell.font          = Font(bold=True, color='FFFFFF')

        # ── Column widths ────────────────────────────────────────────────
        width_map = {
            'name':               28,
            'paye_pin':           18,
            'pension_number':     22,
            'fpa':                20,
            'bank_name':          20,
            'bank_account_number': 20,
        }
        for col_idx, key in enumerate(col_keys, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width_map.get(key, 14)

        # ── Row heights ──────────────────────────────────────────────────
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 36   # wrap text in header

        # ── Save to buffer ───────────────────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _build_pay_map(self, staff) -> tuple[dict[str, Decimal], Decimal, Decimal, Decimal]:
        """
        Return (pay_map, taxable_income, pensionable_income, other_deductions_total)
        for all recurring pay components of a staff.
        taxable_income    = sum of EARNING components where is_taxable=True.
        pensionable_income = sum of EARNING components where is_pensionable=True
                             (Basic Salary + Housing Allowance + Transport Allowance).
        """
        pay_map: dict[str, Decimal] = {}
        taxable_income = Decimal('0')
        pensionable_income = Decimal('0')
        other_deductions_total = Decimal('0')
        for pay_info in staff.pay_info.all():
            comp     = pay_info.component
            name_key = comp.name.lower().strip()
            col_key  = _COMPONENT_NAME_MAP.get(name_key)
            if col_key:
                pay_map[col_key] = pay_map.get(col_key, Decimal('0')) + pay_info.amount
            elif comp.component_type == 'DEDUCTION':
                # Surface deductions that are not part of the legacy fixed template.
                other_deductions_total += pay_info.amount
            if comp.component_type == 'EARNING':
                if comp.is_taxable:
                    taxable_income += pay_info.amount
                if comp.is_pensionable:
                    pensionable_income += pay_info.amount
        return pay_map, taxable_income, pensionable_income, other_deductions_total

    @staticmethod
    def _resolve_payroll_month(period_label: str):
        """Parse month/year labels like 'MARCH 2026' into YYYY-MM-01 date."""
        raw = (period_label or '').strip()
        if not raw:
            now = timezone.now().date()
            return date(now.year, now.month, 1)

        cleaned = re.sub(r'\s+', ' ', raw).strip().title()
        for fmt in ('%B %Y', '%b %Y'):
            try:
                parsed = datetime.strptime(cleaned, fmt).date()
                return date(parsed.year, parsed.month, 1)
            except ValueError:
                continue

        now = timezone.now().date()
        return date(now.year, now.month, 1)

    def _tenant_name(self) -> str:
        staff = self.staff_qs.first()
        if staff and staff.tenant:
            return staff.tenant.name.upper()
        return 'STAFF PAYROLL'

    @staticmethod
    def _default_period_label() -> str:
        now = timezone.now()
        return now.strftime('%B %Y').upper()
