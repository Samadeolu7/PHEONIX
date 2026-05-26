# hr/services/staff_import.py
"""
Staff Payroll Excel Import Service
====================================

Parses an Excel (.xlsx) spreadsheet in the standard payroll upload format
and bulk-creates:
  - Staff records (with bank, pension, and PAYE PIN details)
  - SalaryComponent records (earnings + deductions, if not already present)
  - StaffPayInfo records (linking each staff to their components)

Expected Excel column layout (all other rows before the header are skipped):
  Name | Basic | Housing | Transport | Entertainment | Utility | Lunch |
  Leave Allow | Gross Salary | PAYE Deduct | Loan Deductions |
  Pension Deductions | Dev. Levy & Other | Total Deductions | Net Pay |
  PAYE PIN | PENSION (PEN number) | PFA | Bank | Account Number

Notes:
  - The header row is auto-detected by scanning for the word "Name".
  - Percentage rows (the row immediately after the header showing 16%, 10%…)
    are skipped.
  - Leave Allowance is the ONLY non-taxable component.
  - All other earnings (Basic, Housing, Transport, Entertainment, Utility,
    Lunch) are taxable.
  - PAYE Deduct, Pension Deductions, Loan Deductions, Dev. Levy & Other are
    DEDUCTION components.
  - If a SalaryComponent already exists (by name, branch) it is reused.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Optional

from django.db import transaction as db_transaction

from hr.models import SalaryComponent, Staff, StaffPayInfo

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Column map – canonical header names mapped to internal field keys.
# Matching is case-insensitive and strips whitespace / percentage signs.
# ---------------------------------------------------------------------------

HEADER_ALIASES: dict[str, str] = {
    # Staff identity
    'name': 'name',
    'full name': 'name',
    'staff id': 'staff_id',
    'staff_id': 'staff_id',
    'id': 'staff_id',
    'employee id': 'staff_id',
    'emp id': 'staff_id',
    # Earnings
    'basic': 'basic',
    'basic salary': 'basic',
    'housing': 'housing',
    'housing allowance': 'housing',
    'transport': 'transport',
    'transport allowance': 'transport',
    'entertainment': 'entertainment',
    'entertainment allowance': 'entertainment',
    'utility': 'utility',
    'utility allowance': 'utility',
    'lunch': 'lunch',
    'lunch allowance': 'lunch',
    'leave allow': 'leave_allow',
    'leave allowance': 'leave_allow',
    'leav allow': 'leave_allow',    # abbreviated form used in payroll spreadsheets
    'leav allow.': 'leave_allow',
    'leave allow.': 'leave_allow',
    # Abbreviated entertainment header
    'entertain': 'entertainment',
    'entertain.': 'entertainment',
    # Computed (for validation, not directly stored)
    'gross salary': 'gross_salary',
    'gross': 'gross_salary',
    # Deductions
    'paye deduct': 'paye_deduct',
    'paye deduction': 'paye_deduct',
    'paye': 'paye_deduct',
    'loan deductions': 'loan_deductions',
    'loan': 'loan_deductions',
    'pension deductions': 'pension_deductions',
    'pension': 'pension_deductions',
    'dev. levy & other': 'dev_levy',
    'dev. levy': 'dev_levy',          # split header — "& Other" appears in the row below
    'development levy': 'dev_levy',
    'dev levy': 'dev_levy',
    'dev levy & other': 'dev_levy',
    'total deductions': 'total_deductions',
    'net pay': 'net_pay',
    # Personnel / admin
    'paye pin': 'paye_pin',
    'paye pin number': 'paye_pin',
    'pension number': 'pension_number',
    'pen number': 'pension_number',
    'pension (pen number)': 'pension_number',
    'pfa': 'pfa',
    'pension fund administrator': 'pfa',
    'bank': 'bank_name',
    'bank name': 'bank_name',
    'account number': 'bank_account_number',
    'acct number': 'bank_account_number',
}

# Components to create as EARNING type
EARNING_COMPONENTS: list[dict] = [
    {'key': 'basic',         'name': 'Basic Salary',              'is_taxable': True,  'is_pensionable': True},
    {'key': 'housing',       'name': 'Housing Allowance',         'is_taxable': True,  'is_pensionable': True},
    {'key': 'transport',     'name': 'Transport Allowance',       'is_taxable': True,  'is_pensionable': True},
    {'key': 'entertainment', 'name': 'Entertainment Allowance',   'is_taxable': True,  'is_pensionable': False},
    {'key': 'utility',       'name': 'Utility Allowance',         'is_taxable': True,  'is_pensionable': False},
    {'key': 'lunch',         'name': 'Lunch Allowance',           'is_taxable': True,  'is_pensionable': False},
    # Leave allowance is the ONLY non-taxable component per Nigerian PIT rules;
    # it is also excluded from the pension base.
    {'key': 'leave_allow',   'name': 'Leave Allowance',           'is_taxable': False, 'is_pensionable': False},
]

# PAYE and pension are calculated automatically by the payroll engine;
# importing them from the spreadsheet would cause double-counting.
# Only Loan and Development Levy deductions are taken from the spreadsheet.
DEDUCTION_COMPONENTS: list[dict] = [
    {'key': 'loan_deductions', 'name': 'Loan Deduction'},
    {'key': 'dev_levy',        'name': 'Development Levy'},
]


# ---------------------------------------------------------------------------
# Result dataclasses
# ---------------------------------------------------------------------------

@dataclass
class RowResult:
    row_number: int
    name: str
    status: str          # 'created' | 'updated' | 'skipped' | 'error'
    message: str = ''
    staff_id: Optional[int] = None


@dataclass
class ImportResult:
    total_rows: int = 0
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0
    rows: list[RowResult] = field(default_factory=list)

    @property
    def success(self) -> bool:
        return self.errors == 0


# ---------------------------------------------------------------------------
# Import service
# ---------------------------------------------------------------------------

class StaffImportService:
    """
    Parse a payroll-format Excel workbook and bulk-upsert staff records.

    Usage::

        service = StaffImportService(
            owner=request.user,
            branch=request.user.branch,
            tenant=getattr(request.user, 'tenant', None),
        )
        result  = service.import_from_file(file_obj)
    """

    def __init__(self, owner, branch, tenant=None):
        self.owner  = owner
        self.branch = branch
        self.tenant = tenant

    # ------------------------------------------------------------------
    # Public entry point
    # ------------------------------------------------------------------

    def import_from_file(self, file_obj) -> ImportResult:
        """
        Read an xlsx / xls file-like object and process every data row across
        ALL worksheets in the workbook.

        Each sheet is processed independently; results are accumulated into a
        single ImportResult so that the API response covers all sheets at once.
        Sheets that contain no recognisable 'Name' header are skipped with a
        warning log rather than raising an error (they may be lookup/summary
        tabs).

        Returns an ImportResult summary.
        """
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is required. Install it: pip install openpyxl")

        wb = openpyxl.load_workbook(file_obj, data_only=True)

        result = ImportResult()
        sheets_with_data = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                logger.debug("Sheet '%s' is empty — skipping.", sheet_name)
                continue

            header_row_idx = self._find_header_row(rows)
            if header_row_idx is None:
                logger.warning(
                    "Sheet '%s': could not locate a 'Name' header row — skipping.",
                    sheet_name,
                )
                continue

            headers = rows[header_row_idx]
            col_map = self._build_col_map(headers)

            if 'name' not in col_map:
                logger.warning(
                    "Sheet '%s': header row found but 'Name' column not recognised — skipping.",
                    sheet_name,
                )
                continue

            sheets_with_data += 1

            # ------------------------------------------------------------------
            # Scan ahead from the header row, merging column headers from ALL
            # sub-header rows (percentage row, annotation rows, =N= row, etc.)
            # into col_map before we touch data rows.
            #
            # Payroll spreadsheets commonly have a split-header layout where
            # PAYE PIN / PENSION (PEN number) / PFA / Bank / Account Number
            # labels appear in the =N= currency-indicator row rather than in
            # the main header row.
            #
            # A row is treated as a sub-header row when its Name column is
            # blank or contains a meta placeholder (=N=, -, n/a).
            # ------------------------------------------------------------------
            data_start = header_row_idx + 1
            name_col_idx = col_map.get('name', 0)

            for lookahead_idx in range(data_start, min(data_start + 6, len(rows))):
                candidate = rows[lookahead_idx]
                name_cell = (
                    candidate[name_col_idx]
                    if name_col_idx < len(candidate) else None
                )
                name_cell_str = str(name_cell).strip() if name_cell is not None else ''
                # Real data row found — stop scanning sub-headers.
                if name_cell_str and name_cell_str.lower() not in ('=n=', '-', 'n/a', 'name', ''):
                    break
                # Merge any newly recognised column headers from this sub-header row.
                extra = self._build_col_map(candidate)
                for field_key, col_idx in extra.items():
                    if field_key not in col_map:
                        col_map[field_key] = col_idx
                data_start = lookahead_idx + 1

            # If 'pension_number' was not found via headers — because the
            # PEN-number column often shares the label 'PENSION' with the
            # pension-deductions column — scan the first few data rows for a
            # value that looks like a PEN number (e.g. PEN200673491120).
            if 'pension_number' not in col_map:
                col_map = self._detect_pension_number_col(rows, data_start, col_map)

            for row_idx in range(data_start, len(rows)):
                raw_row = rows[row_idx]
                row_data = self._map_row(raw_row, col_map)

                # Skip completely blank rows
                if not any(v for v in raw_row if v is not None):
                    continue

                name_val = str(row_data.get('name', '')).strip()
                if not name_val:
                    continue

                result.total_rows += 1
                row_result = self._process_row(row_idx + 1, name_val, row_data)
                # Tag result with sheet name for easier debugging
                row_result.message = (
                    f"[Sheet: {sheet_name}] {row_result.message}".strip()
                )
                result.rows.append(row_result)

                if row_result.status == 'created':
                    result.created += 1
                elif row_result.status == 'updated':
                    result.updated += 1
                elif row_result.status == 'skipped':
                    result.skipped += 1
                else:
                    result.errors += 1

        if sheets_with_data == 0:
            raise ValueError(
                "No sheet in the workbook contained a recognisable 'Name' header row. "
                "Ensure the file is in the expected payroll format."
            )

        return result

    # ------------------------------------------------------------------
    # Row processing
    # ------------------------------------------------------------------

    @db_transaction.atomic
    def _process_row(self, row_number: int, full_name: str, row_data: dict) -> RowResult:
        try:
            first_name, last_name = self._split_name(full_name)

            # Extract staff_id from row (e.g. MMA-001, MMC-018)
            raw_staff_id = str(row_data.get('staff_id', '') or '').strip()

            # Lookup order: staff_id first (most specific), then by name
            staff = None
            created = False
            if raw_staff_id and self.tenant:
                try:
                    staff = Staff.objects.get(
                        tenant=self.tenant,
                        branch=self.branch,
                        staff_id=raw_staff_id,
                    )
                except Staff.DoesNotExist:
                    pass

            if staff is None:
                staff, created = Staff.objects.get_or_create(
                    tenant=self.tenant,
                    branch=self.branch,
                    owner=self.owner,
                    first_name=first_name,
                    last_name=last_name,
                    defaults={
                        'staff_id':            raw_staff_id,
                        'paye_pin':            str(row_data.get('paye_pin', '') or '').strip(),
                        'pension_number':      str(row_data.get('pension_number', '') or '').strip(),
                        'pension_provider':    str(row_data.get('pfa', '') or '').strip(),
                        'bank_name':           str(row_data.get('bank_name', '') or '').strip(),
                        'bank_account_number': str(row_data.get('bank_account_number', '') or '').strip(),
                    }
                )

            if not created:
                # Update supplementary fields if they were blank
                changed = False
                for attr, key in [
                    ('staff_id',            'staff_id'),
                    ('paye_pin',            'paye_pin'),
                    ('pension_number',      'pension_number'),
                    ('pension_provider',    'pfa'),
                    ('bank_name',           'bank_name'),
                    ('bank_account_number', 'bank_account_number'),
                ]:
                    new_val = str(row_data.get(key, '') or '').strip()
                    if new_val and not getattr(staff, attr):
                        setattr(staff, attr, new_val)
                        changed = True
                if changed:
                    staff.save()

            # Determine pension exemption from the spreadsheet:
            # If the pension deductions column is blank or zero the staff member
            # is treated as a contract employee (PAYE only, no pension).
            pension_val = self._to_decimal(row_data.get('pension_deductions'))
            is_pension_exempt = (pension_val is None or pension_val == Decimal('0.00'))
            if staff.is_pension_exempt != is_pension_exempt:
                staff.is_pension_exempt = is_pension_exempt
                staff.save(update_fields=['is_pension_exempt'])

            # Ensure components exist and assign to staff
            self._assign_earnings(staff, row_data)
            self._assign_deductions(staff, row_data)

            return RowResult(
                row_number=row_number,
                name=full_name,
                status='created' if created else 'updated',
                staff_id=staff.pk,
            )

        except Exception as exc:
            logger.exception("Error importing row %d (%s): %s", row_number, full_name, exc)
            return RowResult(
                row_number=row_number,
                name=full_name,
                status='error',
                message=str(exc),
            )

    def _assign_earnings(self, staff: Staff, row_data: dict):
        for comp_def in EARNING_COMPONENTS:
            amount = self._to_decimal(row_data.get(comp_def['key']))
            if amount is None:
                continue  # Column not present or blank — skip

            component = self._get_or_create_component(
                name=comp_def['name'],
                component_type=SalaryComponent.EARNING,
                default_amount=amount,
                is_taxable=comp_def['is_taxable'],
                is_pensionable=comp_def['is_pensionable'],
            )
            StaffPayInfo.objects.update_or_create(
                staff=staff,
                component=component,
                defaults={
                    'amount': amount,
                    'branch': self.branch,
                    'owner':  self.owner,
                    'tenant': self.tenant,
                },
            )

    def _assign_deductions(self, staff: Staff, row_data: dict):
        for comp_def in DEDUCTION_COMPONENTS:
            amount = self._to_decimal(row_data.get(comp_def['key']))
            if amount is None or amount == Decimal('0.00'):
                continue  # Skip zero-value deductions

            component = self._get_or_create_component(
                name=comp_def['name'],
                component_type=SalaryComponent.DEDUCTION,
                default_amount=amount,
                is_taxable=False,  # deductions are not taxable
            )
            StaffPayInfo.objects.update_or_create(
                staff=staff,
                component=component,
                defaults={
                    'amount': amount,
                    'branch': self.branch,
                    'owner':  self.owner,
                    'tenant': self.tenant,
                },
            )

    def _get_or_create_component(
        self,
        name: str,
        component_type: str,
        default_amount: Decimal,
        is_taxable: bool = True,
        is_pensionable: bool = False,
    ) -> SalaryComponent:
        """Return existing or newly created SalaryComponent for this branch.

        For existing components the is_taxable and is_pensionable flags are
        updated to stay in sync with the canonical definitions above.
        """
        component, created = SalaryComponent.objects.get_or_create(
            tenant=self.tenant,
            branch=self.branch,
            owner=self.owner,
            name=name,
            defaults={
                'component_type': component_type,
                'default_amount': default_amount,
                'is_taxable':     is_taxable,
                'is_pensionable': is_pensionable,
            },
        )
        if not created:
            # Keep flags in sync for already-existing components.
            changed = False
            if component.is_taxable != is_taxable:
                component.is_taxable = is_taxable
                changed = True
            if component.is_pensionable != is_pensionable:
                component.is_pensionable = is_pensionable
                changed = True
            if changed:
                component.save(update_fields=['is_taxable', 'is_pensionable'])
        return component

    # ------------------------------------------------------------------
    # Header / column detection helpers
    # ------------------------------------------------------------------

    def _find_header_row(self, rows: list) -> Optional[int]:
        """Return the 0-based index of the first row that contains 'Name'."""
        for idx, row in enumerate(rows):
            for cell in row:
                if cell is not None and 'name' in str(cell).strip().lower():
                    return idx
        return None

    def _build_col_map(self, header_row: tuple) -> dict[str, int]:
        """
        Return {field_key: column_index} for recognised headers.
        Unrecognised columns are silently ignored.
        """
        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(header_row):
            if cell is None:
                continue
            # Normalise: lower-case, strip whitespace and percentage signs, collapse spaces
            normalised = str(cell).lower().strip().replace('%', '').strip()
            # Also collapse multiple internal spaces
            normalised = ' '.join(normalised.split())
            if normalised in HEADER_ALIASES:
                field_key = HEADER_ALIASES[normalised]
                # First occurrence wins (handles duplicate-looking headers gracefully)
                if field_key not in col_map:
                    col_map[field_key] = col_idx
        return col_map

    def _detect_pension_number_col(
        self,
        rows: list,
        data_start: int,
        col_map: dict[str, int],
    ) -> dict[str, int]:
        """
        Heuristic fallback: scan the first few data rows for a column whose
        value looks like a PEN number (e.g. PEN200673491120).  This is needed
        because the PEN-number column is often labelled 'PENSION' — the same
        label as the pension-deductions column — so header-based detection
        cannot distinguish them.  The deductions column always wins (first
        occurrence) leaving the PEN-number column unmapped; this method
        finds it from the data instead.
        """
        import re
        pen_re = re.compile(r'^PEN\d{6,}$', re.IGNORECASE)
        already_mapped = set(col_map.values())
        name_col_idx = col_map.get('name', 0)

        for row_idx in range(data_start, min(data_start + 10, len(rows))):
            raw_row = rows[row_idx]
            # Only scan real data rows (non-blank name)
            name_cell = raw_row[name_col_idx] if name_col_idx < len(raw_row) else None
            if not name_cell or not str(name_cell).strip():
                continue
            for col_idx, cell in enumerate(raw_row):
                if col_idx in already_mapped:
                    continue
                if cell is not None and pen_re.match(str(cell).strip()):
                    col_map['pension_number'] = col_idx
                    return col_map
        return col_map

    def _is_percentage_row(self, row: tuple, col_map: dict) -> bool:
        """
        Return True if this row is a percentage/label row that should be skipped.
        Heuristic: the 'basic' or 'housing' column contains a string like '16%'.
        (Kept for compatibility; main scanning now uses the multi-row lookahead.)
        """
        for key in ('basic', 'housing', 'transport'):
            col_idx = col_map.get(key)
            if col_idx is not None and col_idx < len(row):
                val = row[col_idx]
                if val is not None and '%' in str(val):
                    return True
        return False

    def _map_row(self, row: tuple, col_map: dict[str, int]) -> dict:
        """Map raw cell values to field keys using the column map."""
        mapped: dict = {}
        for field_key, col_idx in col_map.items():
            if col_idx < len(row):
                mapped[field_key] = row[col_idx]
        return mapped

    # ------------------------------------------------------------------
    # Utility helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _split_name(full_name: str) -> tuple[str, str]:
        """
        Split 'First Last' into (first_name, last_name).
        If only one word is given, it becomes the last_name.
        For 'Last, First' format, we reverse correctly.
        """
        full_name = full_name.strip()
        if ',' in full_name:
            parts = [p.strip() for p in full_name.split(',', 1)]
            return parts[1], parts[0]
        parts = full_name.split()
        if len(parts) == 1:
            return '', parts[0]
        return parts[0], ' '.join(parts[1:])

    @staticmethod
    def _to_decimal(value) -> Optional[Decimal]:
        """Convert a cell value to Decimal, returning None for blank/non-numeric."""
        if value is None or str(value).strip() in ('', '-', 'N/A', 'n/a'):
            return None
        try:
            cleaned = str(value).replace(',', '').strip()
            d = Decimal(cleaned)
            return d if d >= 0 else None
        except (InvalidOperation, ValueError):
            return None
