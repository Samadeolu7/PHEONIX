# inventory/views_initial_stock_import.py
"""
Initial Inventory Stock Import
===============================
Accepts an Excel workbook and bulk-loads opening stock balances.

Accounting treatment (capital contribution model):
    Dr  Inventory Asset Account  (1124 Trading Stock / per category)
    Cr  Share Capital            (code 3101 - Authorised and Issued Share Capital)

This is used ONCE at system go-live to establish opening stock.
Re-importing the same SKU at the same location is SKIPPED with a warning
to prevent double-posting.

Excel format
------------
Header metadata block (top rows, auto-detected):
    Count Date:   | [date]             |              | Location:    | [name/code]
    Counted By:   | [full name]        |              | Notes:       | [optional]

Data table (header row auto-detected by keyword scanning):
    SKU | Name | Category Code | Category Name | Unit of Measure |
    Opening Qty | Unit Cost | Selling Price | Reorder Level |
    Valuation Method | Barcode | Notes

Standard Category Codes (map to existing FIRS/IFRS COA):
    UNIFORM    → Inventory: 1124, COGS: 5101
    BOOKS      → Inventory: 1124, COGS: 5102
    FOOD       → Inventory: 1124, COGS: 5103
    EVENTS     → Inventory: 1124, COGS: 5104
    TRANSPORT  → Inventory: 1124, COGS: 5105
    STATIONERY → Inventory: 1124, COGS: 5100
    EQUIPMENT  → Inventory: 1124, COGS: 5100
    GENERAL    → Inventory: 1124, COGS: 5100
"""
from __future__ import annotations

import logging
import re as _re
from decimal import Decimal, InvalidOperation
from datetime import date, datetime
from io import BytesIO

from django.db import transaction as db_transaction
from django.utils import timezone
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Standard Category → GL Account mapping
# Keys are the uppercase category codes users will enter in the Excel.
# Values: (inventory_account_code, inventory_account_name, cogs_account_code, cogs_account_name)
# All codes match the FIRS/IFRS COA created by setup_chart_of_accounts.
# ---------------------------------------------------------------------------
STANDARD_CATEGORIES = {
    # All COGS codes are CHILD accounts – never use 5100 (PARENT)
    'UNIFORM':    ('1124', 'Trading Stock / Merchandise', '5101', 'Cost of Sales - Uniform'),
    'BOOKS':      ('1124', 'Trading Stock / Merchandise', '5102', 'Cost of Sales - Textbook'),
    'FOOD':       ('1124', 'Trading Stock / Merchandise', '5103', 'Cost of Sales - Launch'),
    'EVENTS':     ('1124', 'Trading Stock / Merchandise', '5104', 'Cost of Sales - Special Events'),
    'TRANSPORT':  ('1124', 'Trading Stock / Merchandise', '5105', 'Cost of Sales - Transportation'),
    'STATIONERY': ('1124', 'Trading Stock / Merchandise', '5106', 'Cost of Sales - General'),
    'EQUIPMENT':  ('1124', 'Trading Stock / Merchandise', '5106', 'Cost of Sales - General'),
    'LAB':        ('1124', 'Trading Stock / Merchandise', '5106', 'Cost of Sales - General'),
    'GENERAL':    ('1124', 'Trading Stock / Merchandise', '5106', 'Cost of Sales - General'),
    # Raw-material / WIP / finished-goods categories
    'RAW_MAT':    ('1121', 'Raw Materials and Components', '5106', 'Cost of Sales - General'),
    'WIP':        ('1122', 'Work-in-Progress',             '5106', 'Cost of Sales - General'),
    'FINISHED':   ('1123', 'Finished Goods',               '5106', 'Cost of Sales - General'),
}

# Default fallback for custom / unrecognised category codes (both are CHILD accounts)
_DEFAULT_INV_CODE = '1124'
_DEFAULT_INV_NAME = 'Trading Stock / Merchandise'
_DEFAULT_COGS_CODE = '5106'   # Cost of Sales - General (CHILD of 5100)
_DEFAULT_COGS_NAME = 'Cost of Sales - General'

# Opening balance credit account (Share Capital – child posting account)
_SHARE_CAPITAL_CODE = '3101'
_SHARE_CAPITAL_NAME = 'Authorised and Issued Share Capital'

# ---------------------------------------------------------------------------
# Known child-account creation spec
# Keyed by 4-digit code.  Values: (parent_code, child_suffix, name, account_type, parent_name)
# Used by _resolve_account when an account is not yet in the DB (safety net only —
# all standard codes are created by setup_chart_of_accounts).
# ---------------------------------------------------------------------------
_ACCOUNT_CREATE_SPEC = {
    # Inventory asset children (parent 1120 – Inventories)
    '1121': ('1120', '001', 'Raw Materials and Components',    'ASSET',    'Inventories'),
    '1122': ('1120', '002', 'Work-in-Progress',               'ASSET',    'Inventories'),
    '1123': ('1120', '003', 'Finished Goods',                 'ASSET',    'Inventories'),
    '1124': ('1120', '004', 'Trading Stock / Merchandise',    'ASSET',    'Inventories'),
    # Equity children (parent 3100 – Share Capital and Premium)
    '3101': ('3100', '001', 'Authorised and Issued Share Capital', 'EQUITY', 'Share Capital and Premium'),
    # COGS children (parent 5100 – Cost of Sales)
    '5101': ('5100', '001', 'Cost of Sales - Uniform',         'EXPENSE',  'Cost of Sales'),
    '5102': ('5100', '002', 'Cost of Sales - Textbook',        'EXPENSE',  'Cost of Sales'),
    '5103': ('5100', '003', 'Cost of Sales - Launch',          'EXPENSE',  'Cost of Sales'),
    '5104': ('5100', '004', 'Cost of Sales - Special Events',  'EXPENSE',  'Cost of Sales'),
    '5105': ('5100', '005', 'Cost of Sales - Transportation',  'EXPENSE',  'Cost of Sales'),
    '5106': ('5100', '006', 'Cost of Sales - General',         'EXPENSE',  'Cost of Sales'),
}

# Human-readable names shown on InventoryCategory records
_CATEGORY_DISPLAY_NAMES = {
    'UNIFORM':    'School Uniforms',
    'BOOKS':      'Books & Textbooks',
    'FOOD':       'Food & Lunch Items',
    'EVENTS':     'Events & Special Items',
    'TRANSPORT':  'Transport Items',
    'STATIONERY': 'Stationery & Supplies',
    'EQUIPMENT':  'Equipment & Tools',
    'LAB':        'Laboratory Supplies',
    'GENERAL':    'General Inventory',
    'RAW_MAT':    'Raw Materials',
    'WIP':        'Work in Progress',
    'FINISHED':   'Finished Goods',
}

# ── Unicode-aware cell normaliser ────────────────────────────────────────────
_UNICODE_WS = _re.compile(
    r'[\xa0\u00ad\u200b\u200c\u200d\u2060\u202f\u205f\u3000\ufeff]'
)


def _norm(cell_value) -> str:
    """Normalise a cell value to lowercase ASCII-space string."""
    if cell_value is None:
        return ''
    s = _UNICODE_WS.sub(' ', str(cell_value))
    return _re.sub(r'\s+', ' ', s).strip().lower()


def _parse_amount(value) -> Decimal:
    """Convert a cell value to Decimal; return 0 for empty/dash/invalid."""
    if value is None:
        return Decimal('0')
    s = (
        str(value)
        .strip()
        .replace(',', '')
        .replace('₦', '')
        .replace('$', '')
        .replace('£', '')
    )
    if s in ('-', '', 'N/A', 'n/a', '0', 'nil', 'NIL'):
        return Decimal('0')
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal('0')


def _parse_date(value) -> date:
    """Return a Python date from various Excel cell formats."""
    if value is None:
        return date.today()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return date.today()


def _str(value) -> str:
    """Return stripped string or empty string for None."""
    if value is None:
        return ''
    return str(value).strip()


# ── GL account resolver ───────────────────────────────────────────────────────

def _resolve_account(code: str, owner, branch, tenant=None):
    """
    Look up a GL CHILD account by (tenant, branch, code).

    Safety guarantees:
    - Uses (tenant, branch, code) scope — matching how setup_chart_of_accounts
      stores accounts.  Does NOT use owner-based filtering which misses COA entries.
    - If the account is found but is a PARENT (not postable), raises ValueError.
    - If not in the DB, creates it via get_or_create_child_account using
      _ACCOUNT_CREATE_SPEC.  This preserves the parent FK and account_level='CHILD'.
    - Always returns a CHILD / posting account; never a PARENT control account.
    """
    from accounts.models import Account
    from accounts.utils.account_creation import get_or_create_child_account

    # Build scope — (tenant, branch) is the canonical unique key for COA accounts
    scope: dict = {}
    if tenant is not None:
        scope['tenant'] = tenant
    if branch is not None:
        scope['branch'] = branch

    acct = Account.objects.filter(**scope, code=code, is_deleted=False).first()
    if acct is not None:
        if acct.account_level == Account.LEVEL_PARENT:
            raise ValueError(
                f"Account {code} ({acct.name}) is a PARENT/control account and cannot "
                f"be used for direct postings. Only CHILD accounts accept journal entries. "
                f"Check STANDARD_CATEGORIES or _DEFAULT_COGS_CODE in views_initial_stock_import.py."
            )
        return acct

    # Account not in DB — create it via the known spec so the parent FK is set correctly
    if code in _ACCOUNT_CREATE_SPEC:
        parent_code, child_suffix, name, account_type, parent_name = _ACCOUNT_CREATE_SPEC[code]
        return get_or_create_child_account(
            parent_code=parent_code,
            child_suffix=child_suffix,
            name=name,
            account_type=account_type,
            owner=owner,
            branch=branch,
            parent_name=parent_name,
        )

    raise ValueError(
        f"GL account '{code}' was not found for this tenant/branch and is not in the "
        f"known account specification. Run the 'setup_chart_of_accounts' management "
        f"command before importing inventory."
    )


# ── Excel parsing helper ──────────────────────────────────────────────────────

# Keywords used to identify the header row of the data table
_DATA_HEADER_KEYWORDS = {
    # Standard / generic labels
    'sku', 'code', 'item code', 'product code',
    'name', 'description', 'item name', 'product name',
    'category', 'category code', 'category name',
    'uom', 'unit', 'unit of measure', 'unit of measurement',
    'opening qty', 'opening quantity', 'qty', 'quantity',
    'physical count', 'physical qty',
    'unit cost', 'cost', 'cost price',
    'selling price', 'sale price', 'price',
    'reorder level', 'reorder',
    'valuation', 'valuation method', 'method',
    'barcode',
    'notes', 'remarks',
    # Phoenix ERP template hint-row labels (row below the real headers)
    # These allow the parser to work even when the hint row is the only
    # visible header (old template versions or files filled directly).
    'e.g. item-001', 'full item name', 'pick from legend \u25bc',
    'unit / kg / box', 'must be > 0', 'per unit cost',
    'min stock qty', 'average / fifo', 'optional remarks',
}

# Keywords used to identify metadata rows
_META_KEYWORDS = {
    'count date', 'date of count', 'date',
    'location', 'warehouse', 'location/warehouse',
    'counted by', 'counter', 'counter name',
    'notes', 'remarks',
}


def _parse_workbook(wb):
    """
    Parse the workbook.  Returns (metadata, rows, diagnostics).

    metadata: dict with keys count_date, location_name, counted_by, notes
    rows: list of dicts per item
    diagnostics: list of warning strings
    """
    metadata = {
        'count_date': date.today(),
        'location_name': '',
        'counted_by': '',
        'notes': '',
    }
    rows = []
    diagnostics = []

    # Sheet titles that are purely reference/guide sheets — never contain stock data
    _SKIP_TITLES = {'legend', 'instructions', 'instruction', 'reference', 'readme', 'guide', 'notes'}

    for sheet in wb.worksheets:
        if sheet.title.strip().lower() in _SKIP_TITLES:
            continue

        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            continue

        # ── Step 1: extract metadata from the top block ───────────────────────
        # Scan the first 10 rows looking for label → value pairs
        for row in all_rows[:10]:
            for col_idx, cell in enumerate(row):
                # Strip trailing colon/semicolon so "Count Date:" == "Count Date"
                label = _norm(cell).rstrip(':;').strip()
                if label in ('count date', 'date of count', 'date'):
                    # value is in the next non-empty cell
                    for v in row[col_idx + 1:]:
                        if v is not None:
                            metadata['count_date'] = _parse_date(v)
                            break
                elif label in ('location', 'warehouse', 'location/warehouse'):
                    for v in row[col_idx + 1:]:
                        if v is not None:
                            metadata['location_name'] = _str(v)
                            break
                elif label in ('counted by', 'counter name', 'counter'):
                    for v in row[col_idx + 1:]:
                        if v is not None:
                            metadata['counted_by'] = _str(v)
                            break

        # ── Step 2: find the header row of the data table ─────────────────────
        header_row_idx = None
        col_map: dict[str, int] = {}

        for i, row in enumerate(all_rows):
            normalised = [_norm(c) for c in row]
            hits = sum(1 for n in normalised if n in _DATA_HEADER_KEYWORDS)
            # Need at least 3 keyword hits and a name/sku column to be confident
            has_name = any(n in normalised for n in (
                'name', 'description', 'item name', 'product name', 'full item name'
            ))
            has_sku = any(n in normalised for n in (
                'sku', 'code', 'item code', 'product code', 'e.g. item-001'
            ))
            if hits >= 3 and (has_name or has_sku):
                header_row_idx = i
                col_map = {n: j for j, n in enumerate(normalised) if n}
                break

        if header_row_idx is None:
            diagnostics.append(
                f"Sheet '{sheet.title}': no data header row detected — skipped."
            )
            continue

        # ── Step 3: helper to get column by alias ─────────────────────────────
        def _col(*keys):
            for k in keys:
                if k in col_map:
                    return col_map[k]
            return None

        def _get(row, col):
            return row[col] if col is not None and col < len(row) else None

        sku_col = _col('sku', 'code', 'item code', 'product code', 'e.g. item-001')
        name_col = _col('name', 'description', 'item name', 'product name', 'full item name')
        cat_code_col = _col('category code', 'cat code', 'pick from legend \u25bc')
        cat_name_col = _col('category name', 'category')
        uom_col = _col('uom', 'unit of measure', 'unit of measurement', 'unit', 'unit / kg / box')
        qty_col = _col('opening qty', 'opening quantity', 'qty', 'quantity',
                       'physical count', 'physical qty', 'must be > 0')
        cost_col = _col('unit cost', 'cost price', 'cost', 'per unit cost')
        sell_col = _col('selling price', 'sale price', 'price')
        reorder_col = _col('reorder level', 'reorder', 'min stock qty')
        valuation_col = _col('valuation method', 'valuation', 'method', 'average / fifo')
        barcode_col = _col('barcode')
        notes_col = _col('notes', 'remarks', 'optional remarks')
        location_col = _col('location', 'warehouse', 'location/warehouse')

        # ── Step 4: parse data rows ────────────────────────────────────────────
        for row_num, row in enumerate(all_rows[header_row_idx + 1:], start=header_row_idx + 2):
            if all(v is None for v in row):
                continue

            name = _str(_get(row, name_col))
            sku_raw = _str(_get(row, sku_col))

            # Skip obvious non-data rows
            if not name and not sku_raw:
                continue
            low_name = _norm(name)
            low_sku = _norm(sku_raw)
            if low_name in ('name', 'description', 'total', 'subtotal', 'grand total', 's/n', 'sn',
                            'full item name', 'item name', 'item description'):
                continue
            # Skip the template's hint/sub-header row (e.g. ITEM-001 in SKU col)
            if low_sku.startswith('e.g.') or low_sku in ('sku', 'item code', 'product code', 'item sku'):
                continue

            qty = _parse_amount(_get(row, qty_col))
            unit_cost = _parse_amount(_get(row, cost_col))

            # Auto-generate SKU from name if absent
            sku = sku_raw if sku_raw else _re.sub(r'\s+', '-', name.upper())[:50]

            rows.append({
                'row_num': row_num,
                'sku': sku,
                'name': name or sku,
                'category_code': _str(_get(row, cat_code_col)),
                'category_name': _str(_get(row, cat_name_col)),
                'uom': _str(_get(row, uom_col)) or 'unit',
                'opening_qty': qty,
                'unit_cost': unit_cost,
                'selling_price': _parse_amount(_get(row, sell_col)) or unit_cost,
                'reorder_level': _parse_amount(_get(row, reorder_col)),
                'valuation_method': _str(_get(row, valuation_col)).lower() or 'average',
                'barcode': _str(_get(row, barcode_col)),
                'notes': _str(_get(row, notes_col)),
                'location_name': _str(_get(row, location_col)),  # per-row override
                'sheet': sheet.title,
            })

    return metadata, rows, diagnostics


# ── ViewSet ───────────────────────────────────────────────────────────────────

class InitialStockImportViewSet(viewsets.ViewSet):
    """
    Provides a single action:
        POST /api/inventory/initial-stock-import/

    Upload an Excel file to bulk-load opening inventory balances.
    Accounting entry: Dr Inventory Asset / Cr Share Capital.
    """
    permission_classes = [IsAuthenticated]

    @action(
        detail=False,
        methods=['post'],
        url_path='upload',
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload(self, request):
        """
        Accept Excel upload and process initial inventory.
        """
        import openpyxl
        from accounts.models import Account
        from transactions.models import (
            Transaction as JournalEntry,
            TransactionEntry as JournalEntryLine,
            TransactionSeries,
        )
        from .models import (
            InventoryCategory,
            InventoryItem,
            InventoryStock,
            Location,
            StockMovement,
        )

        owner = request.user
        branch = getattr(request.user, 'branch', None)
        tenant = getattr(request, 'tenant', None) or getattr(request.user, 'tenant', None)

        # ── Validate upload ───────────────────────────────────────────────────
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided. Please upload an Excel file (.xlsx or .xls).'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        excel_file = request.FILES['file']
        filename = excel_file.name.lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            return Response(
                {'error': 'File must be an Excel file (.xlsx or .xls)'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Parse workbook ────────────────────────────────────────────────────
        try:
            wb = openpyxl.load_workbook(BytesIO(excel_file.read()), data_only=True)
        except Exception as exc:
            return Response(
                {'error': f'Could not open Excel file: {exc}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        metadata, rows, parse_diagnostics = _parse_workbook(wb)

        if not rows:
            return Response(
                {
                    'error': 'No valid data rows found in the workbook. '
                             'Make sure the spreadsheet has the required columns (SKU, Name, '
                             'Category Code, Opening Qty, Unit Cost).',
                    'diagnostics': parse_diagnostics,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Validate required fields per row ──────────────────────────────────
        validation_errors = []
        for r in rows:
            missing = []
            if not r['name']:
                missing.append('Name')
            if r['opening_qty'] <= 0:
                missing.append('Opening Qty (must be > 0)')
            if r['unit_cost'] <= 0:
                missing.append('Unit Cost (must be > 0)')
            if not r['category_code'] and not r['category_name']:
                missing.append('Category Code or Category Name')
            if missing:
                validation_errors.append(
                    {
                        'row': r['row_num'],
                        'sku': r['sku'],
                        'name': r['name'],
                        'error': f"Missing required field(s): {', '.join(missing)}",
                    }
                )

        # We allow partial import — rows with errors are reported but not processed.
        # Use a set of error row numbers for O(1) lookup — DO NOT compare dict objects.
        _error_row_nums = {ve['row'] for ve in validation_errors}
        valid_rows = [r for r in rows if r['row_num'] not in _error_row_nums]

        # ── Process in one atomic DB transaction ──────────────────────────────
        items_created = 0
        items_existing = 0
        items_skipped_stock = 0
        stock_records_created = 0
        details = []

        try:
            with db_transaction.atomic():
                # ── Ensure Share Capital GL account exists ────────────────────
                share_capital_account = _resolve_account(
                    _SHARE_CAPITAL_CODE, owner, branch, tenant
                )

                # ── Ensure TransactionSeries for opening stock exists ─────────
                series, _ = TransactionSeries.objects.get_or_create(
                    code='INIT',
                    defaults={'description': 'Opening Balance / Initial Stock Entries'},
                )

                # ── Determine default location ────────────────────────────────
                default_location_name = metadata['location_name'] or 'Main Warehouse'
                default_location, loc_created = Location.objects.get_or_create(
                    branch=branch,
                    name=default_location_name,
                    defaults={
                        'code': _re.sub(r'\s+', '_', default_location_name.upper())[:20],
                        'location_type': 'warehouse',
                        'owner': owner,
                    },
                )
                if tenant and not getattr(default_location, 'tenant_id', None):
                    default_location.tenant = tenant
                    default_location.save(update_fields=['tenant'])

                # ── Journal entry accumulators (grouped by inventory account) ─
                # key: inventory_account pk → (account_obj, accumulated_debit)
                debit_accumulator: dict[int, tuple] = {}
                grand_total = Decimal('0')

                # ── Process each valid row ─────────────────────────────────────
                for r in valid_rows:
                    # ─ Resolve location ───────────────────────────────────────
                    loc_name = r['location_name'] or default_location_name
                    if loc_name != default_location_name:
                        location, _ = Location.objects.get_or_create(
                            branch=branch,
                            name=loc_name,
                            defaults={
                                'code': _re.sub(r'\s+', '_', loc_name.upper())[:20],
                                'location_type': 'warehouse',
                                'owner': owner,
                            },
                        )
                        if tenant and not getattr(location, 'tenant_id', None):
                            location.tenant = tenant
                            location.save(update_fields=['tenant'])
                    else:
                        location = default_location

                    # ─ Resolve / auto-create InventoryCategory ────────────────
                    cat_code = (r['category_code'] or r['category_name'] or 'GENERAL').upper().strip()[:20]

                    # Resolve CHILD GL accounts using standard map, fall back to defaults.
                    # IMPORTANT: all codes here must be CHILD accounts, never PARENT.
                    _cat_map = STANDARD_CATEGORIES.get(cat_code)
                    if _cat_map:
                        inv_code, _inv_name, cogs_code, _cogs_name = _cat_map
                    else:
                        inv_code = _DEFAULT_INV_CODE
                        cogs_code = _DEFAULT_COGS_CODE

                    inv_account  = _resolve_account(inv_code,  owner, branch, tenant)
                    cogs_account = _resolve_account(cogs_code, owner, branch, tenant)

                    cat_display_name = r['category_name'] or _CATEGORY_DISPLAY_NAMES.get(cat_code, cat_code)
                    cat_defaults = {
                        'name': cat_display_name,
                        'inventory_account': inv_account,
                        'cogs_account': cogs_account,
                        'owner': owner,
                    }
                    if tenant:
                        cat_defaults['tenant'] = tenant

                    category, cat_created = InventoryCategory.objects.get_or_create(
                        branch=branch,
                        code=cat_code,
                        defaults=cat_defaults,
                    )

                    # ─ Normalise valuation method ─────────────────────────────
                    valuation = r['valuation_method']
                    if valuation not in ('fifo', 'lifo', 'average'):
                        valuation = 'average'

                    # ─ Get or create InventoryItem ────────────────────────────
                    item_defaults = {
                        'name': r['name'],
                        'category': category,
                        'unit_of_measure': r['uom'],
                        'cost_price': r['unit_cost'],
                        'selling_price': r['selling_price'] or r['unit_cost'],
                        'valuation_method': valuation,
                        'reorder_level': r['reorder_level'],
                        'reorder_quantity': Decimal('0'),
                        'is_active': True,
                        'is_sellable': True,
                        'is_purchasable': True,
                        'owner': owner,
                        'barcode': r['barcode'],
                    }
                    if tenant:
                        item_defaults['tenant'] = tenant

                    item, item_created = InventoryItem.objects.get_or_create(
                        branch=branch,
                        sku=r['sku'],
                        defaults=item_defaults,
                    )
                    if item_created:
                        items_created += 1
                    else:
                        items_existing += 1

                    # ─ Check if stock already exists at this location ──────────
                    existing_stock = InventoryStock.objects.filter(
                        item=item, location=location
                    ).first()

                    if existing_stock and existing_stock.quantity_on_hand > 0:
                        # Stock already posted — skip to prevent double-entry
                        items_skipped_stock += 1
                        details.append({
                            'row': r['row_num'],
                            'sku': item.sku,
                            'name': item.name,
                            'status': 'skipped',
                            'reason': f'Already has {existing_stock.quantity_on_hand} units at {location.name}',
                            'qty': str(r['opening_qty']),
                            'unit_cost': str(r['unit_cost']),
                            'total_value': str(r['opening_qty'] * r['unit_cost']),
                        })
                        continue

                    # ─ Create / update InventoryStock ──────────────────────────
                    item_value = r['opening_qty'] * r['unit_cost']

                    if existing_stock:
                        # Stock record exists but quantity is 0 — update it
                        existing_stock.quantity_on_hand = r['opening_qty']
                        existing_stock.quantity_available = r['opening_qty']
                        existing_stock.quantity_reserved = Decimal('0')
                        existing_stock.average_cost = r['unit_cost']
                        existing_stock.total_value = item_value
                        existing_stock.save()
                        stock_record = existing_stock
                    else:
                        stock_defaults = {
                            'quantity_on_hand': r['opening_qty'],
                            'quantity_available': r['opening_qty'],
                            'quantity_reserved': Decimal('0'),
                            'average_cost': r['unit_cost'],
                            'total_value': item_value,
                            'owner': owner,
                        }
                        if tenant:
                            stock_defaults['tenant'] = tenant
                        stock_record = InventoryStock.objects.create(
                            item=item,
                            location=location,
                            branch=branch,
                            **stock_defaults,
                        )
                    stock_records_created += 1

                    # ─ Create StockMovement audit record ──────────────────────
                    movement_defaults = {
                        'movement_type': 'adjustment',
                        'movement_date': metadata['count_date'],
                        'reference_number': f'INIT-{item.sku[:80]}-{date.today().isoformat()}',  # max 100 chars
                        'to_location': location,
                        'quantity': r['opening_qty'],
                        'unit_cost': r['unit_cost'],
                        'total_cost': item_value,
                        'notes': f"Opening balance import. Counted by: {metadata['counted_by'] or 'N/A'}. {r['notes']}".strip('. '),
                        'owner': owner,
                        'branch': branch,
                    }
                    if tenant:
                        movement_defaults['tenant'] = tenant
                    StockMovement.objects.create(item=item, **movement_defaults)

                    # ─ Accumulate journal debit per inventory account ──────────
                    acc_pk = category.inventory_account.pk
                    if acc_pk not in debit_accumulator:
                        debit_accumulator[acc_pk] = (category.inventory_account, Decimal('0'))
                    acc_obj, acc_total = debit_accumulator[acc_pk]
                    debit_accumulator[acc_pk] = (acc_obj, acc_total + item_value)
                    grand_total += item_value

                    details.append({
                        'row': r['row_num'],
                        'sku': item.sku,
                        'name': item.name,
                        'category': category.name,
                        'location': location.name,
                        'status': 'imported',
                        'qty': str(r['opening_qty']),
                        'unit_cost': str(r['unit_cost']),
                        'total_value': str(item_value),
                        'item_created': item_created,
                    })

                # ── Create ONE aggregate journal for the entire batch ──────────
                if grand_total > 0:
                    journal = JournalEntry(
                        series=series,
                        date=metadata['count_date'],
                        description=(
                            f"Opening inventory stock — {stock_records_created} item(s) "
                            f"uploaded by {metadata['counted_by'] or owner.get_full_name() or owner.username}"
                        ),
                        owner=owner,
                        branch=branch,
                    )
                    if tenant:
                        journal.tenant = tenant
                    journal.save()

                    # Debit lines — one per distinct inventory account
                    for acc_obj, acc_debit_total in debit_accumulator.values():
                        JournalEntryLine.objects.create(
                            transaction=journal,
                            account=acc_obj,
                            side=JournalEntryLine.DEBIT,
                            amount=acc_debit_total,
                        )

                    # Credit line — Share Capital
                    JournalEntryLine.objects.create(
                        transaction=journal,
                        account=share_capital_account,
                        side=JournalEntryLine.CREDIT,
                        amount=grand_total,
                    )

                    # Post the journal (updates account balances)
                    journal.post()

        except Exception as exc:
            logger.exception('initial_stock_import: unexpected error during processing')
            return Response(
                {'error': f'Processing failed: {str(exc)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response(
            {
                'success': True,
                'summary': {
                    'total_rows_parsed': len(rows),
                    'items_created': items_created,
                    'items_existing_updated': items_existing - items_skipped_stock,
                    'stock_records_posted': stock_records_created,
                    'items_skipped_existing_stock': items_skipped_stock,
                    'validation_errors': len(validation_errors),
                    'grand_total_value': str(grand_total),
                    'count_date': metadata['count_date'].isoformat(),
                    'location': metadata['location_name'] or 'Main Warehouse',
                    'counted_by': metadata['counted_by'],
                    'journal_posted': grand_total > 0,
                    'accounting_entry': (
                        'Dr Trading Stock / Inventory (1124) / Cr Share Capital (3101)'
                        if grand_total > 0
                        else 'No journal (nothing to post)'
                    ),
                },
                'parse_diagnostics': parse_diagnostics,
                'validation_errors': validation_errors,
                'details': details,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=['get'], url_path='generate-template')
    def generate_template(self, request):
        """
        Generate and return a .xlsx template with:
        - Sheet "Instructions": usage guide
        - Sheet "Stock Upload": data entry with Category Code dropdown + auto-fill name
        - Sheet "LEGEND": category code reference with GL account mapping
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        LEGEND_DATA = [
            # (code, display_name, inv_account, cogs_account, description)
            ('UNIFORM',    'School Uniforms',         '1124 - Trading Stock', '5101 - COS Uniform',        'Clothing items sold to students'),
            ('BOOKS',      'Books & Textbooks',       '1124 - Trading Stock', '5102 - COS Textbook',       'Textbooks and educational materials'),
            ('FOOD',       'Food & Lunch Items',      '1124 - Trading Stock', '5103 - COS Launch',         'Canteen / tuck-shop consumables'),
            ('EVENTS',     'Events & Special Items',  '1124 - Trading Stock', '5104 - COS Special Events', 'Items for school events and ceremonies'),
            ('TRANSPORT',  'Transport Items',         '1124 - Trading Stock', '5105 - COS Transportation', 'Fuel, spare parts, transport supplies'),
            ('STATIONERY', 'Stationery & Supplies',   '1124 - Trading Stock', '5106 - COS General',        'Office and classroom stationery'),
            ('EQUIPMENT',  'Equipment & Tools',       '1124 - Trading Stock', '5106 - COS General',        'Small tools and equipment (not fixed assets)'),
            ('LAB',        'Laboratory Supplies',     '1124 - Trading Stock', '5106 - COS General',        'Science lab consumables and reagents'),
            ('GENERAL',    'General Inventory',       '1124 - Trading Stock', '5106 - COS General',        'Catch-all for items not in the above categories'),
            ('RAW_MAT',    'Raw Materials',           '1121 - Raw Materials', '5106 - COS General',        'Production / manufacturing raw materials'),
            ('WIP',        'Work in Progress',        '1122 - Work-in-Progress', '5106 - COS General',     'Partially completed goods'),
            ('FINISHED',   'Finished Goods',          '1123 - Finished Goods', '5106 - COS General',       'Completed manufactured products'),
        ]
        NUM_LEGEND = len(LEGEND_DATA)

        wb = openpyxl.Workbook()

        # ── Colour palette ────────────────────────────────────────────────────
        BLUE_DARK   = 'FF1F3864'
        BLUE_MID    = 'FF2E74B5'
        BLUE_LIGHT  = 'FFDBE5F1'
        GREY_LIGHT  = 'FFF2F2F2'
        YELLOW      = 'FFFFF2CC'
        GREEN_LIGHT = 'FFE2EFDA'
        ORANGE_SOFT = 'FFFCE4D6'
        WHITE       = 'FFFFFFFF'

        # helper: fill
        def fill(hex_colour):
            return PatternFill(fill_type='solid', fgColor=hex_colour)

        # helper: border
        thin = Side(style='thin', color='FF999999')
        medium = Side(style='medium', color='FF444444')

        def thin_border():
            return Border(left=thin, right=thin, top=thin, bottom=thin)

        def bottom_border():
            return Border(bottom=Side(style='medium', color='FF1F3864'))

        # ── SHEET 1: LEGEND ───────────────────────────────────────────────────
        ws_legend = wb.active
        ws_legend.title = 'LEGEND'

        # Header
        ws_legend.column_dimensions['A'].width = 14
        ws_legend.column_dimensions['B'].width = 30
        ws_legend.column_dimensions['C'].width = 32
        ws_legend.column_dimensions['D'].width = 32
        ws_legend.column_dimensions['E'].width = 45

        legend_headers = ['Category Code', 'Category Name', 'Inventory Account', 'COGS Account', 'Description']
        for col_idx, hdr in enumerate(legend_headers, start=1):
            cell = ws_legend.cell(row=1, column=col_idx, value=hdr)
            cell.font = Font(bold=True, color='FFFFFFFF', size=11)
            cell.fill = fill(BLUE_DARK)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border()
        ws_legend.row_dimensions[1].height = 20

        # Data rows
        for row_idx, (code, name, inv_acc, cogs_acc, desc) in enumerate(LEGEND_DATA, start=2):
            row_fill = fill(GREY_LIGHT) if row_idx % 2 == 0 else fill(WHITE)
            data = [code, name, inv_acc, cogs_acc, desc]
            for col_idx, val in enumerate(data, start=1):
                cell = ws_legend.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = row_fill
                cell.border = thin_border()
                cell.alignment = Alignment(vertical='center', wrap_text=(col_idx == 5))
                if col_idx == 1:
                    cell.font = Font(bold=True, color='FF1F3864')
            ws_legend.row_dimensions[row_idx].height = 16

        # Title row above headers
        ws_legend.insert_rows(1)
        ws_legend.merge_cells('A1:E1')
        t = ws_legend['A1']
        t.value = 'INVENTORY CATEGORY LEGEND — Reference for Category Code column in Stock Upload sheet'
        t.font = Font(bold=True, size=12, color='FFFFFFFF')
        t.fill = fill(BLUE_MID)
        t.alignment = Alignment(horizontal='center', vertical='center')
        ws_legend.row_dimensions[1].height = 24

        # ── SHEET 2: Stock Upload ─────────────────────────────────────────────
        ws = wb.create_sheet('Stock Upload')

        # Column widths
        col_widths = {
            'A': 5,   # SN
            'B': 18,  # SKU
            'C': 32,  # Name
            'D': 16,  # Category Code  (dropdown)
            'E': 28,  # Category Name  (VLOOKUP — auto-fill)
            'F': 16,  # Unit of Measure
            'G': 14,  # Opening Qty
            'H': 14,  # Unit Cost
            'I': 14,  # Selling Price
            'J': 14,  # Reorder Level
            'K': 18,  # Valuation Method
            'L': 18,  # Barcode
            'M': 30,  # Notes
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # ── Metadata block (rows 1-5) ─────────────────────────────────────────
        # Row 1: branding banner
        ws.merge_cells('A1:M1')
        banner = ws['A1']
        banner.value = 'PHOENIX ERP — Initial Inventory Opening Balance Upload Template'
        banner.font = Font(bold=True, size=14, color='FFFFFFFF')
        banner.fill = fill(BLUE_DARK)
        banner.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        # Row 2: Count Date + Location
        for c in ['A2', 'B2', 'C2', 'D2', 'E2']:
            ws[c].fill = fill(GREY_LIGHT)
        ws['A2'].value = 'Count Date:'
        ws['A2'].font = Font(bold=True)
        ws['A2'].alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells('B2:C2')
        ws['B2'].font = Font(color='FF1F3864')
        ws['B2'].fill = fill(YELLOW)
        ws['B2'].border = bottom_border()
        ws['B2'].alignment = Alignment(vertical='center')

        ws['F2'].value = 'Location:'
        ws['F2'].font = Font(bold=True)
        ws['F2'].alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells('G2:I2')
        ws['G2'].font = Font(color='FF1F3864')
        ws['G2'].fill = fill(YELLOW)
        ws['G2'].border = bottom_border()
        ws['G2'].alignment = Alignment(vertical='center')
        ws.row_dimensions[2].height = 20

        # Row 3: Counted By + Notes
        ws['A3'].value = 'Counted By:'
        ws['A3'].font = Font(bold=True)
        ws['A3'].alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells('B3:C3')
        ws['B3'].fill = fill(YELLOW)
        ws['B3'].border = bottom_border()
        ws['B3'].alignment = Alignment(vertical='center')

        ws['F3'].value = 'Notes:'
        ws['F3'].font = Font(bold=True)
        ws['F3'].alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells('G3:M3')
        ws['G3'].fill = fill(YELLOW)
        ws['G3'].border = bottom_border()
        ws['G3'].alignment = Alignment(vertical='center')
        ws.row_dimensions[3].height = 20

        # Row 4: blank spacer
        ws.row_dimensions[4].height = 6

        # ── Data table header (row 5) ─────────────────────────────────────────
        HEADERS = [
            ('SN',              'A', False),
            ('SKU',             'B', True),
            ('Name',            'C', True),
            ('Category Code',   'D', True),
            ('Category Name',   'E', False),   # auto-fill
            ('Unit of Measure', 'F', False),
            ('Opening Qty',     'G', True),
            ('Unit Cost',       'H', True),
            ('Selling Price',   'I', False),
            ('Reorder Level',   'J', False),
            ('Valuation Method','K', False),
            ('Barcode',         'L', False),
            ('Notes',           'M', False),
        ]
        DATA_START_ROW = 7
        NUM_DATA_ROWS = 500

        for col_idx, (hdr, col_letter, required) in enumerate(HEADERS, start=1):
            cell = ws.cell(row=5, column=col_idx, value=hdr)
            cell.font = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill = fill(BLUE_DARK)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border()
        ws.row_dimensions[5].height = 30

        # Sub-header hints row (row 5 supplements — add a note row)
        HINTS = [
            '', 'e.g. ITEM-001', 'Full item name', 'Pick from LEGEND ▼', '⟵ Auto-fills',
            'unit / kg / box', 'Must be > 0', 'Per unit cost', 'Sale price', 'Min stock qty',
            'average / fifo', 'Optional', 'Optional remarks',
        ]
        # We'll add hints as a second header row
        for col_idx, hint in enumerate(HINTS, start=1):
            cell = ws.cell(row=DATA_START_ROW - 1, column=col_idx, value=hint)
            cell.font = Font(italic=True, size=8, color='FF444444')
            cell.fill = fill(BLUE_LIGHT)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border()
        ws.row_dimensions[DATA_START_ROW - 1].height = 20

        # Freeze panes: keep banner + header visible
        ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=1)

        # ── Data rows: formulas and formatting ────────────────────────────────
        # After insert_rows(1): title=row1, headers=row2, data=row3..(2+NUM_LEGEND)
        # Codes in col A, Names in col B
        dv = DataValidation(
            type='list',
            formula1=f"LEGEND!$A$3:$A${2 + NUM_LEGEND}",
            allow_blank=True,
            showDropDown=False,    # False = show the dropdown arrow in Excel
            showErrorMessage=True,
            error='Please select a valid category from the dropdown or check the LEGEND sheet.',
            errorTitle='Invalid Category Code',
            showInputMessage=True,
            prompt='Select a category code. See the LEGEND sheet for all options.',
            promptTitle='Category Code',
        )
        ws.add_data_validation(dv)
        dv.sqref = f'D{DATA_START_ROW}:D{DATA_START_ROW + NUM_DATA_ROWS - 1}'

        for row in range(DATA_START_ROW, DATA_START_ROW + NUM_DATA_ROWS):
            is_sample = (row == DATA_START_ROW)

            # SN: auto-number
            ws.cell(row=row, column=1).value = row - DATA_START_ROW + 1
            ws.cell(row=row, column=1).font = Font(size=9, color='FF888888')
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

            # Category Name: VLOOKUP → auto-fill from LEGEND col B given code in col A
            cat_name_cell = ws.cell(row=row, column=5)
            cat_name_cell.value = (
                f'=IF(D{row}="","",IFERROR(VLOOKUP(D{row},LEGEND!$A:$B,2,FALSE),"[Custom — will be auto-created]"))'
            )
            cat_name_cell.font = Font(italic=True, color='FF1F3864', size=9)
            cat_name_cell.fill = fill(BLUE_LIGHT)
            cat_name_cell.alignment = Alignment(vertical='center')

            # Row banding
            row_fill = fill(GREY_LIGHT) if row % 2 == 0 else fill(WHITE)

            for col_idx in range(1, 14):
                cell = ws.cell(row=row, column=col_idx)
                # Highlight required fields in yellow for empty rows
                if col_idx in (2, 3, 4, 7, 8):   # SKU, Name, Cat Code, Qty, Cost
                    if not is_sample:
                        cell.fill = fill(YELLOW) if not cell.value else row_fill
                elif col_idx == 5:
                    pass  # already styled
                else:
                    if not cell.value:
                        cell.fill = row_fill
                cell.border = thin_border()
                cell.alignment = Alignment(vertical='center')

            if is_sample:
                # Pre-fill a sample row so users see the format
                sample_data = {2: 'ITEM-001', 3: 'Example Product Name', 4: 'GENERAL',
                               6: 'unit', 7: 10, 8: 500, 9: 600, 10: 5, 11: 'average'}
                for col_idx, val in sample_data.items():
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = val
                    cell.font = Font(italic=True, color='FF888888')
                    cell.fill = fill(GREY_LIGHT)

        # Re-style required header cells to stand out
        for col_idx, (hdr, col_letter, required) in enumerate(HEADERS, start=1):
            if required:
                ws.cell(row=5, column=col_idx).fill = fill(BLUE_MID)

        # ── SHEET 3: Instructions ─────────────────────────────────────────────
        ws_inst = wb.create_sheet('Instructions')
        ws_inst.column_dimensions['A'].width = 3
        ws_inst.column_dimensions['B'].width = 80

        inst_rows = [
            (None, None),
            ('title',    'PHOENIX ERP — Initial Inventory Upload: Instructions'),
            (None, None),
            ('h2',       'PURPOSE'),
            ('body',     'Use this template to upload your opening inventory balances into Phoenix ERP.'),
            ('body',     'This is a one-time operation performed during system go-live.'),
            (None, None),
            ('h2',       'STEPS'),
            ('step',     '1.  Open the "Stock Upload" sheet.'),
            ('step',     '2.  Fill in Count Date, Location, and Counted By in the yellow cells at the top.'),
            ('step',     '3.  For each item, enter SKU, Name, and select a Category Code from the dropdown.'),
            ('step',     '4.  The Category Name column fills in automatically — do NOT type in it.'),
            ('step',     '5.  Enter Opening Qty (must be > 0) and Unit Cost (must be > 0).'),
            ('step',     '6.  Save the file and upload via:  Inventory → Initial Stock Import.'),
            (None, None),
            ('h2',       'COLUMN GUIDE'),
            ('required', 'SKU             — Unique item code. Leave blank to auto-generate from Name (not recommended).'),
            ('required', 'Name            — Full item description.'),
            ('required', 'Category Code   — Select from the dropdown (see LEGEND sheet for all options).'),
            ('auto',     'Category Name   — Fills automatically from the LEGEND. Do not type here.'),
            ('optional', 'Unit of Measure — unit, kg, litre, box, pack, etc. Defaults to "unit".'),
            ('required', 'Opening Qty     — Physical count quantity. Must be greater than zero.'),
            ('required', 'Unit Cost       — Cost price per unit at the time of upload.'),
            ('optional', 'Selling Price   — Defaults to Unit Cost if left blank.'),
            ('optional', 'Reorder Level   — Minimum stock before a replenishment alert is triggered.'),
            ('optional', 'Valuation Method— average, fifo, or lifo. Defaults to "average".'),
            ('optional', 'Barcode         — For barcode scanners. Optional.'),
            ('optional', 'Notes           — Any item-level remarks.'),
            (None, None),
            ('h2',       'ACCOUNTING ENTRY'),
            ('body',     'For each uploaded item the system posts:'),
            ('body',     '    Dr  1124 — Trading Stock / Merchandise  (Inventory Asset)'),
            ('body',     '    Cr  3101 — Authorised and Issued Share Capital'),
            ('body',     'One aggregate journal entry is created per upload batch.'),
            (None, None),
            ('h2',       'SAFETY RULES'),
            ('body',     '• Rows where the SKU already has stock at that location are SKIPPED (no double-posting).'),
            ('body',     '• Custom category codes (not in the LEGEND) are accepted — they map to account 1124 / 5100.'),
            ('body',     '• The entire upload is wrapped in one database transaction — errors roll everything back.'),
            ('body',     '• Multiple locations per batch are supported via the per-row "Location" column (if added).'),
        ]

        for i, (style, text) in enumerate(inst_rows, start=1):
            ws_inst.row_dimensions[i].height = 16
            if style is None:
                continue
            cell = ws_inst.cell(row=i, column=2, value=text)
            if style == 'title':
                cell.font = Font(bold=True, size=16, color='FFFFFFFF')
                cell.fill = fill(BLUE_DARK)
                ws_inst.merge_cells(f'A{i}:B{i}')
                ws_inst['A' + str(i)].fill = fill(BLUE_DARK)
                ws_inst.row_dimensions[i].height = 28
            elif style == 'h2':
                cell.font = Font(bold=True, size=11, color='FFFFFFFF')
                cell.fill = fill(BLUE_MID)
                ws_inst.merge_cells(f'A{i}:B{i}')
                ws_inst['A' + str(i)].fill = fill(BLUE_MID)
                ws_inst.row_dimensions[i].height = 20
            elif style == 'step':
                cell.font = Font(size=10, bold=True)
            elif style == 'required':
                cell.font = Font(size=10, color='FF1F3864')
                # small required dot
                ws_inst.cell(row=i, column=1, value='●').font = Font(color='FFFF0000', size=7)
            elif style == 'auto':
                cell.font = Font(size=10, color='FF2E74B5', italic=True)
                ws_inst.cell(row=i, column=1, value='⟵').font = Font(color='FF2E74B5', size=8)
            elif style == 'optional':
                cell.font = Font(size=10, color='FF444444')
            else:
                cell.font = Font(size=10)
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        # Widen col B
        ws_inst.column_dimensions['B'].width = 90

        # Reorder sheets: Instructions → Stock Upload → LEGEND
        wb.move_sheet('Instructions', offset=-2)  # move to front
        wb.move_sheet('Stock Upload', offset=-1)

        # ── Stream response ───────────────────────────────────────────────────
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="phoenix_initial_stock_template.xlsx"'
        return response

    @action(detail=False, methods=['get'], url_path='template')
    def template(self, request):
        """Redirect callers to the proper generate-template endpoint."""
        return self.generate_template(request)

