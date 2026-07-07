# banks/parsers.py
"""
Bank statement parsers for the Bank-Recon integration.

Parses uploaded bank statement files (CSV, Excel, QIF) into a normalised
list of transaction dicts that Django forwards to the Java
IngestAndMatchController. Use ``parse_statement_file`` to dispatch on the
uploaded file's extension rather than instantiating a parser directly.

Tabular formats (CSV / Excel) share the same column layout assumption
(a header row containing at least a date column and a debit and/or credit
column — column position and any leading metadata rows do not matter):
  Transaction Date | Narration | Reference | Debit | Credit | Balance

Rules applied to every format:
  - Amounts have comma-thousands separators: "50,000.00" -> Decimal('50000.00')
  - Direction is inferred from the sign (QIF) or from which of
    Debit/Credit is populated (CSV/Excel)
  - A hash of (account_number, date, amount, direction, narration) is used as
    bank_ref when no reference/cheque number is supplied (idempotency key
    for Java's dedup-on-reupload logic)
"""

import csv
import hashlib
import io
import logging
import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import IO, List, Optional

logger = logging.getLogger(__name__)


@dataclass
class ParsedTransaction:
    bank_ref: str          # unique reference (from file or hash-derived)
    value_date: str        # ISO-8601 string: YYYY-MM-DD
    narration: str
    direction: str         # 'CREDIT' | 'DEBIT'
    amount: str            # Decimal string, e.g. "50000.00"
    balance_after: str     # Decimal string or empty string


class _TabularStatementParser:
    """
    Shared column-detection and row-parsing logic for statement formats that
    are fundamentally a grid of rows/columns (CSV, Excel). Subclasses only
    need to turn their source format into a list of rows (each row a list
    of stripped strings) and call ``_build_transactions``.
    """

    # Column header synonyms (case-insensitive, after strip)
    _COL_DATE       = {'transaction date', 'date', 'trans date', 'value date'}
    _COL_NARRATION  = {'narration', 'description', 'transaction description'}
    _COL_REFERENCE  = {'reference', 'ref', 'cheque no', 'instrument no'}
    _COL_DEBIT      = {'debit', 'withdrawal', 'dr'}
    _COL_CREDIT     = {'credit', 'deposit', 'cr'}
    _COL_BALANCE    = {'balance', 'running balance', 'ledger balance'}

    @classmethod
    def _build_transactions(cls, rows: List[List[str]], account_number: str) -> List[ParsedTransaction]:
        if not rows:
            raise ValueError("Statement file is empty.")

        header_idx, col_map = cls._find_header(rows)
        if header_idx is None:
            raise ValueError(
                "Could not locate header row. Expected columns: "
                "Transaction Date, Narration, Debit, Credit."
            )

        transactions: List[ParsedTransaction] = []
        for row_num, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            if not any(cell.strip() for cell in row):
                continue  # blank row

            try:
                txn = cls._parse_row(row, col_map, account_number, row_num)
            except (ValueError, InvalidOperation) as exc:
                logger.warning("Row %d skipped — %s", row_num, exc)
                continue

            if txn:
                transactions.append(txn)

        if not transactions:
            raise ValueError(
                "No valid transactions found in the statement. "
                "Please verify the file format."
            )

        return transactions

    # ── internals ────────────────────────────────────────────────────────────

    @classmethod
    def _find_header(cls, rows):
        """Return (header_row_index, column_map) or (None, None)."""
        for idx, row in enumerate(rows):
            normalised = [c.strip().lower() for c in row]
            col_map = {}
            for col_idx, name in enumerate(normalised):
                if name in cls._COL_DATE:
                    col_map['date'] = col_idx
                elif name in cls._COL_NARRATION:
                    col_map['narration'] = col_idx
                elif name in cls._COL_REFERENCE:
                    col_map['reference'] = col_idx
                elif name in cls._COL_DEBIT:
                    col_map['debit'] = col_idx
                elif name in cls._COL_CREDIT:
                    col_map['credit'] = col_idx
                elif name in cls._COL_BALANCE:
                    col_map['balance'] = col_idx
            if 'date' in col_map and ('debit' in col_map or 'credit' in col_map):
                return idx, col_map
        return None, None

    @classmethod
    def _parse_row(cls, row, col_map, account_number, row_num):
        """Parse one data row; return ParsedTransaction or None if row should be skipped."""

        def get(key, default=''):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return default
            return row[idx].strip()

        raw_date      = get('date')
        raw_narration = get('narration')
        raw_reference = get('reference')
        raw_debit     = get('debit')
        raw_credit    = get('credit')
        raw_balance   = get('balance')

        # Parse date
        value_date = cls._parse_date(raw_date)
        if value_date is None:
            if not raw_date:
                return None  # likely a sub-total / footer row
            raise ValueError(f"Unrecognised date format: '{raw_date}'")

        # Parse amounts
        debit_amount  = cls._parse_amount(raw_debit)
        credit_amount = cls._parse_amount(raw_credit)
        balance       = cls._parse_amount(raw_balance)

        if debit_amount is None and credit_amount is None:
            return None  # no amount — skip

        if credit_amount is not None and credit_amount > Decimal('0'):
            direction = 'CREDIT'
            amount = credit_amount
        elif debit_amount is not None and debit_amount > Decimal('0'):
            direction = 'DEBIT'
            amount = debit_amount
        else:
            return None  # both zero — skip

        bank_ref = cls._derive_bank_ref(
            raw_reference, account_number, value_date, amount, direction, raw_narration
        )

        return ParsedTransaction(
            bank_ref=bank_ref,
            value_date=value_date,
            narration=raw_narration[:500],
            direction=direction,
            amount=str(amount),
            balance_after=str(balance) if balance is not None else '',
        )

    @staticmethod
    def _derive_bank_ref(raw_reference, account_number, value_date, amount, direction, narration) -> str:
        if raw_reference:
            return raw_reference[:200]
        # Deterministic hash so Java can deduplicate on re-upload
        h = hashlib.sha256(
            f"{account_number}|{value_date}|{amount}|{direction}|{narration}"
            .encode('utf-8')
        ).hexdigest()[:32]
        return f"HASH-{h}"

    @staticmethod
    def _parse_date(raw: str) -> Optional[str]:
        """Return ISO date string or None."""
        if not raw:
            return None
        raw = raw.strip()
        for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y'):
            try:
                return datetime.strptime(raw, fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_amount(raw: str) -> Optional[Decimal]:
        """Strip commas and return Decimal, or None for blank/dash."""
        if not raw or raw.strip() in ('', '-', 'N/A'):
            return None
        cleaned = raw.replace(',', '').replace(' ', '')
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return None


class FirstBankStatementParser(_TabularStatementParser):
    """
    Parses a CSV bank statement export (First Bank of Nigeria layout, but
    any CSV with a Date/Narration/Reference/Debit/Credit/Balance header
    works regardless of bank).

    Usage::

        with open('statement.csv', 'rb') as fh:
            transactions = FirstBankStatementParser.parse(fh, account_number='1234567890')
    """

    @classmethod
    def parse(cls, file_obj: IO[bytes], account_number: str = '') -> List[ParsedTransaction]:
        """
        Parse a CSV file object (binary mode) and return a list of ParsedTransaction.

        Raises ValueError with a human-readable message if the file cannot be parsed.
        """
        try:
            text = file_obj.read()
            if isinstance(text, bytes):
                # Try UTF-8 first, fall back to latin-1 (common in bank exports)
                try:
                    text = text.decode('utf-8-sig')  # strip BOM if present
                except UnicodeDecodeError:
                    text = text.decode('latin-1')
        except Exception as exc:
            raise ValueError(f"Cannot read statement file: {exc}") from exc

        rows = list(csv.reader(io.StringIO(text)))
        return cls._build_transactions(rows, account_number)


class ExcelStatementParser(_TabularStatementParser):
    """
    Parses an .xlsx bank statement export (e.g. Moniepoint account
    statement download). Reuses the same header-name detection as
    FirstBankStatementParser — leading metadata rows (account name,
    opening/closing balance, etc.) and non-contiguous columns (from merged
    narration cells) are handled the same way a sparse CSV row would be.

    Usage::

        with open('statement.xlsx', 'rb') as fh:
            transactions = ExcelStatementParser.parse(fh, account_number='1234567890')
    """

    @classmethod
    def parse(cls, file_obj: IO[bytes], account_number: str = '') -> List[ParsedTransaction]:
        try:
            import openpyxl
        except ImportError as exc:
            raise ValueError(
                "Excel statement support requires the 'openpyxl' package."
            ) from exc

        try:
            raw = file_obj.read()
            workbook = cls._load_workbook_tolerant(raw, openpyxl)
            worksheet = workbook[workbook.sheetnames[0]]
            rows = [
                [cls._cell_to_str(cell) for cell in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
        except ValueError:
            raise
        except Exception as exc:
            raise ValueError(f"Cannot read Excel statement: {exc}") from exc

        return cls._build_transactions(rows, account_number)

    @staticmethod
    def _cell_to_str(value) -> str:
        if value is None:
            return ''
        if isinstance(value, (datetime, date)):
            return value.strftime('%d/%m/%Y')
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)

    @staticmethod
    def _load_workbook_tolerant(raw: bytes, openpyxl):
        """
        Some bank statement exporters (observed on Moniepoint downloads) emit
        an .xlsx with two issues real Excel would never produce:
          1. styles.xml uses invalid casing for alignment enum values (e.g.
             vertical="Top" instead of the OOXML-spec "top"), which openpyxl
             rejects outright with "could not read stylesheet".
          2. The worksheet's <dimension ref="A1"/> tag understates the real
             data range, which makes openpyxl's read_only mode (which trusts
             that tag) see only the first cell. Normal mode ignores the
             dimension tag and parses the full <sheetData>, so read_only is
             deliberately NOT used here despite the larger memory footprint.
        Rather than fail the whole upload over cosmetic/incorrect XML we
        don't actually need, patch the casing in memory and retry.
        """
        try:
            return openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        except ValueError as exc:
            if 'stylesheet' not in str(exc).lower():
                raise

        zin = zipfile.ZipFile(io.BytesIO(raw))
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == 'xl/styles.xml':
                    text = data.decode('utf-8')
                    text = re.sub(
                        r'(vertical|horizontal)="([A-Za-z]+)"',
                        lambda m: f'{m.group(1)}="{m.group(2).lower()}"',
                        text,
                    )
                    data = text.encode('utf-8')
                zout.writestr(item, data)
        buf.seek(0)
        return openpyxl.load_workbook(buf, data_only=True)


class QifStatementParser:
    """
    Parses a Quicken Interchange Format (QIF) bank statement export.

    QIF record fields used here:
      D  transaction date
      T  amount (signed: positive = credit/inflow, negative = debit/outflow)
      N  reference / transaction number
      P  payee
      M  memo
    Records are separated by a line containing only '^'. Any other field
    code (L=category, A=address, etc.) is ignored — a bank export carries no
    such fields anyway.

    Usage::

        with open('statement.qif', 'rb') as fh:
            transactions = QifStatementParser.parse(fh, account_number='1234567890')
    """

    @classmethod
    def parse(cls, file_obj: IO[bytes], account_number: str = '') -> List[ParsedTransaction]:
        try:
            text = file_obj.read()
            if isinstance(text, bytes):
                try:
                    text = text.decode('utf-8-sig')
                except UnicodeDecodeError:
                    text = text.decode('latin-1')
        except Exception as exc:
            raise ValueError(f"Cannot read statement file: {exc}") from exc

        lines = [line.strip() for line in text.splitlines() if not line.strip().startswith('!')]

        transactions: List[ParsedTransaction] = []
        record: dict = {}
        record_num = 0

        for line in lines:
            if not line:
                continue
            if line == '^':
                record_num += 1
                txn = cls._flush(record, account_number, record_num)
                if txn:
                    transactions.append(txn)
                record = {}
                continue

            code, value = line[0], line[1:]
            if code == 'D':
                record['date'] = value
            elif code == 'T':
                record['amount'] = value
            elif code == 'N':
                record['reference'] = value
            elif code == 'P':
                record['payee'] = value
            elif code == 'M':
                record['memo'] = value

        if record:  # trailing record without a final '^' — malformed but recoverable
            record_num += 1
            txn = cls._flush(record, account_number, record_num)
            if txn:
                transactions.append(txn)

        if not transactions:
            raise ValueError(
                "No valid transactions found in the QIF file. "
                "Please verify the file format."
            )

        return transactions

    # ── internals ────────────────────────────────────────────────────────────

    @classmethod
    def _flush(cls, record: dict, account_number: str, record_num: int) -> Optional[ParsedTransaction]:
        if not record:
            return None
        try:
            return cls._build_transaction(record, account_number)
        except (ValueError, InvalidOperation) as exc:
            logger.warning("QIF record %d skipped — %s", record_num, exc)
            return None

    @classmethod
    def _build_transaction(cls, record: dict, account_number: str) -> ParsedTransaction:
        raw_date = record.get('date', '')
        value_date = cls._parse_qif_date(raw_date)
        if value_date is None:
            raise ValueError(f"Unrecognised date format: '{raw_date}'")

        raw_amount = record.get('amount', '')
        signed_amount = cls._parse_signed_amount(raw_amount)
        if signed_amount is None:
            raise ValueError(f"Unrecognised amount: '{raw_amount}'")
        if signed_amount == 0:
            raise ValueError("Zero-amount record")

        direction = 'CREDIT' if signed_amount > 0 else 'DEBIT'
        amount = abs(signed_amount)

        narration = (record.get('memo') or record.get('payee') or '').strip()[:500]

        raw_reference = (record.get('reference') or '').strip()
        if raw_reference:
            bank_ref = raw_reference[:200]
        else:
            h = hashlib.sha256(
                f"{account_number}|{value_date}|{amount}|{direction}|{narration}"
                .encode('utf-8')
            ).hexdigest()[:32]
            bank_ref = f"HASH-{h}"

        return ParsedTransaction(
            bank_ref=bank_ref,
            value_date=value_date,
            narration=narration,
            direction=direction,
            amount=str(amount),
            balance_after='',
        )

    @staticmethod
    def _parse_qif_date(raw: str) -> Optional[str]:
        if not raw:
            return None
        raw = raw.strip().replace(' ', '')
        # Some QIF exports use an apostrophe for the century, e.g. "7/1'26" -> "7/1/2026"
        if "'" in raw:
            head, _, year = raw.partition("'")
            raw = f"{head}/20{year}" if len(year) == 2 else f"{head}/{year}"
        for fmt in ('%d/%m/%Y', '%d/%m/%y', '%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d', '%d-%m-%Y'):
            try:
                return datetime.strptime(raw, fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_signed_amount(raw: str) -> Optional[Decimal]:
        if raw is None:
            return None
        cleaned = raw.strip().replace(',', '').replace(' ', '')
        if not cleaned:
            return None
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return None


_EXTENSION_PARSERS = {
    '.qif': QifStatementParser,
    '.xlsx': ExcelStatementParser,
    '.xlsm': ExcelStatementParser,
}


def parse_statement_file(file_obj: IO[bytes], filename: str, account_number: str = '') -> List[ParsedTransaction]:
    """
    Dispatch to the right parser based on the uploaded file's extension.
    Defaults to the CSV parser for '.csv', '.txt', or any unrecognised
    extension, since that has always been the primary supported format.
    """
    ext = ''
    if filename:
        dot = filename.rfind('.')
        if dot != -1:
            ext = filename[dot:].lower()
    parser = _EXTENSION_PARSERS.get(ext, FirstBankStatementParser)
    return parser.parse(file_obj, account_number=account_number)
